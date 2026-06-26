"""Danh gia voice_test_N_MFCC theo Chunk ID."""
import argparse
import os
import sys

from sound.config import DEFAULT_XLSX, guided_test_phases

PHASES_DEFAULT = guided_test_phases()


def expected_for_chunk(chunk_id, phases):
    for label, command, start, end in phases:
        if start <= chunk_id <= end:
            return label, command
    return '', ''


def evaluate_sheet(ws, phases):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    def cell(row, name, default=''):
        idx = col.get(name)
        if not idx:
            return default
        v = ws.cell(row, idx).value
        return v if v is not None else default

    stats = {'total_rows': 0, 'too_quiet': 0, 'mfcc_cmds': 0, 'mfcc_reject': 0, 'by_phase': {}}
    for label, command, _, _ in phases:
        stats['by_phase'][label] = {
            'command': command, 'rows': 0, 'cmd_rows': 0,
            'Y': 0, 'N': 0, 'correct_cmd': 0, 'wrong_cmd': 0,
        }

    for row in range(2, ws.max_row + 1):
        try:
            chunk_id = int(cell(row, 'Chunk ID', 0))
        except (TypeError, ValueError):
            continue
        exp_label, exp_cmd = expected_for_chunk(chunk_id, phases)
        if not exp_label:
            continue
        stats['total_rows'] += 1
        note = str(cell(row, 'Ghi chu', ''))
        received = str(cell(row, 'Lenh nhan', '-'))
        phase = stats['by_phase'][exp_label]
        phase['rows'] += 1
        if note == 'too_quiet':
            stats['too_quiet'] += 1
        if note.startswith('mfcc_reject'):
            stats['mfcc_reject'] += 1
        if received not in ('-', ''):
            stats['mfcc_cmds'] += 1
            phase['cmd_rows'] += 1
            if received == exp_cmd:
                phase['Y'] += 1
            else:
                phase['N'] += 1
    return stats


def print_report(test_no, sheet_name, stats, phases):
    print('')
    print('=== voice_test_%d_MFCC (%s) ===' % (test_no, sheet_name))
    for label, command, start, end in phases:
        print('  %s (%s): chunk %d-%d' % (label, command, start, end))
    print('')
    print('Tong dong trong phase: %d' % stats['total_rows'])
    print('too_quiet: %d (%.1f%%)' % (
        stats['too_quiet'], 100.0 * stats['too_quiet'] / max(stats['total_rows'], 1)))
    print('mfcc lenh fire: %d | mfcc_reject: %d' % (stats['mfcc_cmds'], stats['mfcc_reject']))
    total_y = total_n = 0
    for label, command, _, _ in phases:
        p = stats['by_phase'][label]
        scored = p['Y'] + p['N']
        total_y += p['Y']
        total_n += p['N']
        acc = 100.0 * p['Y'] / scored if scored else 0.0
        print('%s (%s): %d dong | %d lenh | Y=%d N=%d | acc=%.1f%%' % (
            label, command, p['rows'], p['cmd_rows'], p['Y'], p['N'], acc))
    scored_all = total_y + total_n
    if scored_all:
        print('')
        print('Accuracy tong (Y/(Y+N)): %.1f%% (%d/%d)' % (
            100.0 * total_y / scored_all, total_y, scored_all))


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument('--test', type=int, default=6)
    parser.add_argument('--xlsx', default='')
    args = parser.parse_args(argv)
    xlsx = args.xlsx or DEFAULT_XLSX
    sheet = 'voice_test_%d_MFCC' % args.test
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('pip install openpyxl')
        sys.exit(1)
    if not os.path.isfile(xlsx):
        print('Khong tim thay:', xlsx)
        sys.exit(1)
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        print('Khong co sheet:', sheet)
        print('Co:', ', '.join(wb.sheetnames))
        sys.exit(1)
    stats = evaluate_sheet(wb[sheet], PHASES_DEFAULT)
    print_report(args.test, sheet, stats, PHASES_DEFAULT)
    wb.close()


if __name__ == '__main__':
    main()
