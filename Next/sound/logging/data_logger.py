import os
import re
import threading
import time

from sound.config import DEFAULT_XLSX

_DEFAULT_HEADERS = (
    'STT', 'Phien test', 'Thoi gian', 'Chunk ID', 'Am mong muon', 'Text STT',
    'Lenh mong muon', 'Lenh nhan', 'Dung/Sai', 'Latency STT (ms)',
    'Latency action (ms)', 'Phuong phap', 'Ghi chu',
)
_TEST_SHEET_RE = re.compile(r'^voice_test_(\d+)_(MFCC|STT)$', re.I)


class VoiceDataLogger(object):
    def __init__(self, xlsx_path=None, stt_backend='', test_mode='STT'):
        env_path = os.environ.get('MARIO_VOICE_XLSX', '').strip()
        self.xlsx_path = xlsx_path or env_path or DEFAULT_XLSX
        self.stt_backend = stt_backend
        self.test_mode = (test_mode or 'STT').upper()
        if self.test_mode not in ('MFCC', 'STT'):
            self.test_mode = 'STT'
        self.session_id = time.strftime('%Y%m%d_%H%M%S')
        self._lock = threading.Lock()
        self._row_no = 0
        self._enabled = False
        self._wb = None
        self._ws = None
        self._summary_ws = None
        self._rows_written = 0
        self.test_no = 0
        self.log_sheet_name = ''
        self.summary_sheet_name = ''
        self._setup_workbook()

    @staticmethod
    def _next_test_number(workbook, test_mode):
        test_mode = test_mode.upper()
        max_no = 0
        for name in workbook.sheetnames:
            m = _TEST_SHEET_RE.match(name)
            if m and m.group(2).upper() == test_mode:
                max_no = max(max_no, int(m.group(1)))
        return max_no + 1

    @staticmethod
    def _unique_sheet_name(workbook, base_name):
        name = base_name[:31]
        if name not in workbook.sheetnames:
            return name
        for i in range(2, 100):
            suffix = '_%d' % i
            candidate = (base_name[:31 - len(suffix)] + suffix)
            if candidate not in workbook.sheetnames:
                return candidate
        return base_name[:28] + '_x'

    def _create_log_sheet(self, workbook):
        self.test_no = self._next_test_number(workbook, self.test_mode)
        self.log_sheet_name = self._unique_sheet_name(
            workbook, 'voice_test_%d_%s' % (self.test_no, self.test_mode))
        self.summary_sheet_name = self._unique_sheet_name(
            workbook, '%s_summary' % self.log_sheet_name)
        ws = workbook.create_sheet(self.log_sheet_name)
        for col, title in enumerate(_DEFAULT_HEADERS, start=1):
            ws.cell(1, col, title)
        summary = workbook.create_sheet(self.summary_sheet_name)
        summary.append(('Metric', 'Value'))
        summary.append(('Test sheet', self.log_sheet_name))
        summary.append(('Phuong phap', self.test_mode))
        summary.append(('STT backend', self.stt_backend))
        summary.append(('Session', self.session_id))
        return ws

    def _setup_workbook(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            print('[VOICE] Excel log disabled. pip install openpyxl')
            return
        try:
            if os.path.isfile(self.xlsx_path):
                self._wb = load_workbook(self.xlsx_path)
            else:
                self._wb = Workbook()
                default = self._wb.active
                default.title = 'Sheet1'
            self._ws = self._create_log_sheet(self._wb)
            self._summary_ws = self._wb[self.summary_sheet_name]
            self._row_no = 0
            self._wb.save(self.xlsx_path)
            self._enabled = True
            print('[VOICE] Excel log:', self.xlsx_path)
            print('[VOICE] Sheet moi:', self.log_sheet_name)
            print('[VOICE] Summary:', self.summary_sheet_name)
            print('[VOICE] Phuong phap:', self.test_mode, '| session:', self.session_id)
            print('[VOICE] Dong EXCEL truoc khi choi.')
        except PermissionError:
            print('[VOICE] Excel log failed: file dang mo. Hay DONG Excel roi chay lai.')
            self._enabled = False
        except Exception as exc:
            print('[VOICE] Excel log failed:', exc)
            self._enabled = False

    def log_voice_event(self, text='', command='', chunk_id=0,
                        stt_latency_ms='', action_latency_ms='', note='',
                        expected_label='', expected_command=''):
        if not self._enabled:
            return
        received = command if command else '-'
        expected = expected_command or '-'
        correct = '-'
        if expected_command and received not in ('-', ''):
            correct = 'Y' if received == expected_command else 'N'
        elif expected_command and note in ('stt_empty', 'too_quiet', 'no_command'):
            correct = 'N'
        with self._lock:
            try:
                self._row_no += 1
                row = self._ws.max_row + 1
                values = (
                    self._row_no, self.session_id, time.strftime('%Y-%m-%d %H:%M:%S'),
                    chunk_id, expected_label or '-', text, expected, received, correct,
                    stt_latency_ms, action_latency_ms, self.test_mode, note,
                )
                for col, value in enumerate(values, start=1):
                    self._ws.cell(row, col, value)
                self._wb.save(self.xlsx_path)
                self._rows_written += 1
                if self._rows_written == 1:
                    print('[VOICE] Excel: da ghi dong dau tien vao', self.log_sheet_name)
            except PermissionError:
                print('[VOICE] Excel write error: file dang mo trong Excel. Hay dong file.')
            except Exception as exc:
                print('[VOICE] Excel write error:', exc)

    def log_summary(self, metrics):
        if not self._enabled or self._summary_ws is None:
            return
        with self._lock:
            try:
                ws = self._summary_ws
                ws.append(('Saved at', time.strftime('%Y-%m-%d %H:%M:%S')))
                ws.append(('Rows logged', self._rows_written))
                for key, value in metrics.items():
                    if isinstance(value, list):
                        avg = int(sum(value) / len(value)) if value else 0
                        ws.append((key + ' avg', avg))
                        ws.append((key + ' count', len(value)))
                    else:
                        ws.append((key, value))
                self._wb.save(self.xlsx_path)
            except Exception as exc:
                print('[VOICE] Excel summary error:', exc)

    def close(self):
        if self._enabled:
            print('[VOICE] Excel: %d dong -> %s | %s' % (
                self._rows_written, self.log_sheet_name, self.summary_sheet_name))
        self._wb = None
        self._ws = None
        self._summary_ws = None
        self._enabled = False
